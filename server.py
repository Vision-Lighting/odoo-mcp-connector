"""Odoo MCP Server for Claude Desktop.

Modes (set via ODOO_MODE env var):
  staging  — full access, staging database
  live_ro  — read-only, live database
  live_rw  — restricted writes (products, project tasks, helpdesk tickets), live database
"""

import json
import os
import traceback
from typing import Any

from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent

from odoo_client import client, ODOO_MODE, ODOO_URL, ODOO_DB
from pdfmonkey_client import generate_document, PDFMONKEY_TEMPLATE_ID

# ── Mode configuration ────────────────────────────────────────────────────────

MODE = ODOO_MODE  # staging | live_ro | live_rw

# Models that live_rw may WRITE to
LIVE_RW_WRITE_MODELS = {
    "product.template",
    "product.product",
    "project.task",
    "project.project",
    "product.pricelist.item",
    "helpdesk.ticket",
}
# Models that live_rw may CREATE in
LIVE_RW_CREATE_MODELS = {
    "project.task",
    "sale.order",
    "product.pricelist.item",
    "helpdesk.ticket",
}

MODE_LABELS = {
    "staging":  "🧪 STAGING",
    "live_ro":  "🔵 LIVE (read-only)",
    "live_rw":  "🟠 LIVE (restricted write)",
}

app = Server(f"odoo-{MODE}")


def _ok(data: Any) -> list[TextContent]:
    return [TextContent(type="text", text=json.dumps(data, default=str, indent=2))]


def _err(msg: str) -> list[TextContent]:
    return [TextContent(type="text", text=f"ERROR: {msg}")]


def _denied(action: str) -> list[TextContent]:
    return _err(
        f"Action '{action}' is not permitted in {MODE_LABELS[MODE]} mode. "
        f"Switch to the staging server to perform this operation."
    )


# ── URL helpers ───────────────────────────────────────────────────────────────

# Odoo 18 clean URL patterns (falls back to /web# for unknown models)
_MODEL_URL_PATHS: dict[str, str] = {
    "product.template":      "/odoo/inventory/products/{id}",
    "product.product":       "/odoo/inventory/products/{id}",
    "sale.order":            "/odoo/sales/{id}",
    "purchase.order":        "/odoo/purchase/{id}",
    "project.task":          "/odoo/project/tasks/{id}",
    "project.project":       "/odoo/project/{id}",
    "stock.picking":         "/odoo/inventory/delivery-orders/{id}",
    "res.partner":           "/odoo/contacts/{id}",
    "account.move":          "/odoo/accounting/customer-invoices/{id}",
    "helpdesk.ticket":       "/odoo/helpdesk/tickets/{id}",
}


def record_url(model: str, record_id: int) -> str:
    pattern = _MODEL_URL_PATHS.get(model)
    if pattern:
        return f"{ODOO_URL}{pattern.format(id=record_id)}"
    return f"{ODOO_URL}/web#model={model}&id={record_id}&view_type=form"


def _with_url(record: dict, model: str) -> dict:
    """Add a _url field to a record dict if it has an id."""
    if isinstance(record, dict) and "id" in record:
        record["_url"] = record_url(model, record["id"])
    return record


def _inject_urls(records: list[dict], model: str) -> list[dict]:
    """Add _url to every record in a list."""
    return [_with_url(r, model) for r in records]


# ── Tool catalogue ────────────────────────────────────────────────────────────

# --- Read tools (all modes) ---
READ_TOOLS = [
    Tool(
        name="odoo_ping",
        description=(
            f"Test the Odoo connection. Current mode: {MODE_LABELS[MODE]} | "
            f"DB: {ODOO_DB} | URL: {ODOO_URL}"
        ),
        inputSchema={"type": "object", "properties": {}},
    ),
    Tool(
        name="odoo_search_read",
        description=(
            f"[{MODE_LABELS[MODE]}] Search and read records from any Odoo model."
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "model": {"type": "string", "description": "Odoo model, e.g. 'product.template'"},
                "domain": {"type": "array", "description": "Filter domain, e.g. [['name','ilike','LED']]", "default": []},
                "fields": {"type": "array", "items": {"type": "string"}, "description": "Fields to return (omit for all)"},
                "limit": {"type": "integer", "default": 80},
                "offset": {"type": "integer", "default": 0},
                "order": {"type": "string", "description": "Sort order, e.g. 'name asc'"},
            },
            "required": ["model"],
        },
    ),
    Tool(
        name="odoo_read",
        description=f"[{MODE_LABELS[MODE]}] Read specific Odoo records by ID.",
        inputSchema={
            "type": "object",
            "properties": {
                "model": {"type": "string"},
                "ids": {"type": "array", "items": {"type": "integer"}},
                "fields": {"type": "array", "items": {"type": "string"}},
            },
            "required": ["model", "ids"],
        },
    ),
    Tool(
        name="odoo_get_fields",
        description=(
            f"[{MODE_LABELS[MODE]}] Get field definitions for an Odoo model, "
            "including custom Studio fields (x_ prefix)."
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "model": {"type": "string"},
                "attributes": {"type": "array", "items": {"type": "string"},
                               "description": "e.g. ['string','type','required']"},
            },
            "required": ["model"],
        },
    ),
    Tool(
        name="product_get",
        description=(
            f"[{MODE_LABELS[MODE]}] Get full product template details including "
            "variants, attributes, and custom Studio fields."
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "id": {"type": "integer", "description": "product.template ID"},
                "include_variants": {"type": "boolean", "default": True},
            },
            "required": ["id"],
        },
    ),
    Tool(
        name="sales_find_order",
        description=f"[{MODE_LABELS[MODE]}] Find sales orders by reference or customer name.",
        inputSchema={
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Order ref, e.g. 'S00123'"},
                "customer": {"type": "string", "description": "Customer name (partial)"},
                "state": {"type": "string", "enum": ["draft", "sent", "sale", "done", "cancel"]},
                "limit": {"type": "integer", "default": 20},
            },
        },
    ),
    Tool(
        name="sales_order_get",
        description=f"[{MODE_LABELS[MODE]}] Get a sales order with lines and delivery pickings.",
        inputSchema={
            "type": "object",
            "properties": {
                "id": {"type": "integer", "description": "sale.order ID"},
            },
            "required": ["id"],
        },
    ),
]

# --- Write tools (staging + live_rw) ---
WRITE_TOOLS = [
    Tool(
        name="odoo_create",
        description=(
            f"[{MODE_LABELS[MODE]}] Create a new record. "
            + ("Allowed models: " + ", ".join(sorted(LIVE_RW_CREATE_MODELS))
               if MODE == "live_rw" else "Any model.")
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "model": {"type": "string"},
                "values": {"type": "object"},
            },
            "required": ["model", "values"],
        },
    ),
    Tool(
        name="odoo_write",
        description=(
            f"[{MODE_LABELS[MODE]}] Update Odoo records. "
            + ("Allowed models: " + ", ".join(sorted(LIVE_RW_WRITE_MODELS))
               if MODE == "live_rw" else "Any model.")
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "model": {"type": "string"},
                "ids": {"type": "array", "items": {"type": "integer"}},
                "values": {"type": "object"},
            },
            "required": ["model", "ids", "values"],
        },
    ),
    Tool(
        name="product_create",
        description=f"[{MODE_LABELS[MODE]}] Create a new product template.",
        inputSchema={
            "type": "object",
            "properties": {
                "name": {"type": "string"},
                "product_type": {"type": "string", "enum": ["consu", "service", "product"], "default": "product"},
                "sale_price": {"type": "number"},
                "cost_price": {"type": "number"},
                "internal_reference": {"type": "string"},
                "description": {"type": "string"},
                "description_sale": {"type": "string"},
                "categ_id": {"type": "integer"},
                "extra_fields": {"type": "object", "description": "Any additional/Studio fields"},
            },
            "required": ["name"],
        },
    ),
    Tool(
        name="product_set_image",
        description=f"[{MODE_LABELS[MODE]}] Set image on a product template or variant.",
        inputSchema={
            "type": "object",
            "properties": {
                "model": {"type": "string", "enum": ["product.template", "product.product"], "default": "product.template"},
                "id": {"type": "integer"},
                "image_url": {"type": "string"},
                "image_base64": {"type": "string"},
            },
            "required": ["id"],
        },
    ),
    Tool(
        name="product_update_variant",
        description=(
            f"[{MODE_LABELS[MODE]}] Update fields on a product variant, "
            "including custom Studio fields (x_ prefix)."
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "variant_id": {"type": "integer"},
                "values": {"type": "object"},
            },
            "required": ["variant_id", "values"],
        },
    ),
    Tool(
        name="task_upsert",
        description=(
            f"[{MODE_LABELS[MODE]}] Create or update a project task. "
            "If task_id is provided, updates it; otherwise creates a new task."
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "task_id": {"type": "integer", "description": "Existing task ID to update (omit to create)"},
                "name": {"type": "string", "description": "Task title"},
                "project_id": {"type": "integer", "description": "Project ID"},
                "description": {"type": "string", "description": "Task description / notes"},
                "user_ids": {"type": "array", "items": {"type": "integer"}, "description": "Assigned user IDs"},
                "stage_id": {"type": "integer", "description": "Stage/column ID"},
                "date_deadline": {"type": "string", "description": "Due date, e.g. '2025-06-30'"},
                "priority": {"type": "string", "enum": ["0", "1"], "description": "0=normal, 1=high"},
                "tag_ids": {"type": "array", "items": {"type": "integer"}, "description": "Tag IDs"},
                "extra_fields": {"type": "object", "description": "Any additional/Studio fields"},
            },
        },
    ),
    Tool(
        name="helpdesk_ticket_upsert",
        description=(
            f"[{MODE_LABELS[MODE]}] Create or update a helpdesk ticket. "
            "If ticket_id is provided, updates it; otherwise creates a new ticket."
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "ticket_id": {"type": "integer", "description": "Existing ticket ID to update (omit to create)"},
                "name": {"type": "string", "description": "Ticket subject/title"},
                "partner_id": {"type": "integer", "description": "Customer (res.partner) ID"},
                "partner_name": {"type": "string", "description": "Customer name (used if partner_id unknown)"},
                "description": {"type": "string", "description": "Ticket description / body"},
                "team_id": {"type": "integer", "description": "Helpdesk team ID"},
                "user_id": {"type": "integer", "description": "Assigned agent user ID"},
                "stage_id": {"type": "integer", "description": "Stage ID"},
                "priority": {"type": "string", "enum": ["0", "1", "2", "3"], "description": "0=low, 1=medium, 2=high, 3=urgent"},
                "tag_ids": {"type": "array", "items": {"type": "integer"}, "description": "Tag IDs"},
                "extra_fields": {"type": "object", "description": "Any additional/Studio fields"},
            },
        },
    ),
]

# --- Datasheet tool (staging + live_ro + live_rw — read-only operation) ---
DATASHEET_TOOLS = [
    Tool(
        name="generate_datasheet",
        description=(
            f"[{MODE_LABELS[MODE]}] Generate a PDF datasheet for a product variant via PDFMonkey. "
            "Fetches all spec fields and the product image from Odoo, builds the payload, "
            "submits to PDFMonkey, and returns the download URL. "
            "Accepts a variant internal reference (SKU), a product.product ID, or a product name to search."
        ),
        inputSchema={
            "type": "object",
            "properties": {
                "internal_reference": {
                    "type": "string",
                    "description": "Product variant SKU / internal reference (default_code), e.g. 'PLB-1230-24W-4K-9-MP'",
                },
                "variant_id": {
                    "type": "integer",
                    "description": "product.product ID (use instead of internal_reference if you already have it)",
                },
                "product_name": {
                    "type": "string",
                    "description": "Partial product name to search (returns error if multiple matches found)",
                },
                "template_id": {
                    "type": "string",
                    "description": f"PDFMonkey template ID override. Defaults to {PDFMONKEY_TEMPLATE_ID}",
                },
            },
        },
    ),
]

# --- Staging-only tools ---
STAGING_ONLY_TOOLS = [
    Tool(
        name="odoo_call",
        description="[🧪 STAGING] Call any method on an Odoo model.",
        inputSchema={
            "type": "object",
            "properties": {
                "model": {"type": "string"},
                "method": {"type": "string"},
                "ids": {"type": "array", "items": {"type": "integer"}},
                "kwargs": {"type": "object", "default": {}},
            },
            "required": ["model", "method", "ids"],
        },
    ),
    Tool(
        name="delivery_split",
        description="[🧪 STAGING] Split a delivery picking into multiple batches.",
        inputSchema={
            "type": "object",
            "properties": {
                "picking_id": {"type": "integer"},
                "batches": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "scheduled_date": {"type": "string"},
                            "moves": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "move_id": {"type": "integer"},
                                        "qty": {"type": "number"},
                                    },
                                    "required": ["move_id", "qty"],
                                },
                            },
                        },
                        "required": ["moves"],
                    },
                },
            },
            "required": ["picking_id", "batches"],
        },
    ),
    Tool(
        name="delivery_split_from_excel",
        description="[🧪 STAGING] Read a delivery schedule from Excel and auto-split a picking.",
        inputSchema={
            "type": "object",
            "properties": {
                "picking_id": {"type": "integer"},
                "excel_path": {"type": "string"},
                "sheet_name": {"type": "string"},
                "product_col": {"type": "string", "default": "Product"},
                "qty_col": {"type": "string", "default": "Qty"},
                "date_col": {"type": "string", "default": "Date"},
            },
            "required": ["picking_id", "excel_path"],
        },
    ),
]

# Build final tool list for this mode
if MODE == "staging":
    ALL_TOOLS = READ_TOOLS + WRITE_TOOLS + DATASHEET_TOOLS + STAGING_ONLY_TOOLS
elif MODE == "live_rw":
    ALL_TOOLS = READ_TOOLS + WRITE_TOOLS + DATASHEET_TOOLS
else:  # live_ro
    ALL_TOOLS = READ_TOOLS + DATASHEET_TOOLS


# ── Handlers ──────────────────────────────────────────────────────────────────

@app.list_tools()
async def list_tools() -> list[Tool]:
    return ALL_TOOLS


@app.call_tool()
async def call_tool(name: str, arguments: dict) -> list[TextContent]:
    try:
        # Block tools not in this mode's list
        allowed = {t.name for t in ALL_TOOLS}
        if name not in allowed:
            return _denied(name)
        return await _dispatch(name, arguments)
    except Exception as e:
        return _err(f"{e}\n\n{traceback.format_exc()}")


async def _dispatch(name: str, args: dict) -> list[TextContent]:

    # ── Ping ──────────────────────────────────────────────────────────────────
    if name == "odoo_ping":
        import xmlrpc.client as xc
        common = xc.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
        version = common.version()
        uid = client.uid
        user = client.read("res.users", [uid], ["name", "login", "company_id"])
        return _ok({
            "mode": MODE_LABELS[MODE],
            "server_version": version,
            "uid": uid,
            "user": user,
            "db": ODOO_DB,
            "url": ODOO_URL,
        })

    # ── Read tools ────────────────────────────────────────────────────────────
    if name == "odoo_search_read":
        model = args["model"]
        records = client.search_read(
            model,
            domain=args.get("domain", []),
            fields=args.get("fields"),
            limit=args.get("limit", 80),
            offset=args.get("offset", 0),
            order=args.get("order"),
        )
        return _ok(_inject_urls(records, model))

    if name == "odoo_read":
        model = args["model"]
        records = client.read(model, args["ids"], fields=args.get("fields"))
        return _ok(_inject_urls(records, model))

    if name == "odoo_get_fields":
        return _ok(client.get_fields(args["model"], attributes=args.get("attributes")))

    if name == "product_get":
        tmpl = _with_url(client.get_product_template(args["id"]), "product.template")
        result: dict = {"template": tmpl}
        if args.get("include_variants", True):
            result["variants"] = _inject_urls(
                client.get_product_variants(args["id"]), "product.product"
            )
        return _ok(result)

    if name == "sales_find_order":
        domain: list = []
        if "name" in args:
            domain.append(["name", "ilike", args["name"]])
        if "customer" in args:
            domain.append(["partner_id.name", "ilike", args["customer"]])
        if "state" in args:
            domain.append(["state", "=", args["state"]])
        records = client.search_read(
            "sale.order", domain,
            fields=["id", "name", "partner_id", "state", "date_order", "amount_total", "picking_ids"],
            limit=args.get("limit", 20),
            order="date_order desc",
        )
        return _ok(_inject_urls(records, "sale.order"))

    if name == "sales_order_get":
        order = _with_url(client.get_sale_order(args["id"]), "sale.order")
        lines = client.search_read(
            "sale.order.line", [["order_id", "=", args["id"]]],
            fields=["id", "product_id", "product_uom_qty", "qty_delivered", "price_unit", "name"],
        )
        pickings = client.get_pickings_for_sale_order(args["id"])
        for p in pickings:
            p["moves"] = client.get_move_lines(p["id"])
            _with_url(p, "stock.picking")
        return _ok({"order": order, "lines": lines, "pickings": pickings})

    # ── Write tools ───────────────────────────────────────────────────────────
    if name == "odoo_create":
        model = args["model"]
        if MODE == "live_rw" and model not in LIVE_RW_CREATE_MODELS:
            return _err(
                f"Cannot create '{model}' in {MODE_LABELS[MODE]} mode. "
                f"Allowed: {', '.join(sorted(LIVE_RW_CREATE_MODELS))}"
            )
        new_id = client.create(model, args["values"])
        return _ok({"id": new_id, "model": model, "url": record_url(model, new_id)})

    if name == "odoo_write":
        model = args["model"]
        if MODE == "live_rw" and model not in LIVE_RW_WRITE_MODELS:
            return _err(
                f"Cannot write to '{model}' in {MODE_LABELS[MODE]} mode. "
                f"Allowed: {', '.join(sorted(LIVE_RW_WRITE_MODELS))}"
            )
        ok = client.write(model, args["ids"], args["values"])
        return _ok({
            "success": ok,
            "ids": args["ids"],
            "urls": [record_url(model, i) for i in args["ids"]],
        })

    if name == "product_create":
        values: dict = {"name": args["name"]}
        if "product_type" in args:
            values["type"] = args["product_type"]
        if "sale_price" in args:
            values["list_price"] = args["sale_price"]
        if "internal_reference" in args:
            values["default_code"] = args["internal_reference"]
        if "description" in args:
            values["description"] = args["description"]
        if "description_sale" in args:
            values["description_sale"] = args["description_sale"]
        if "categ_id" in args:
            values["categ_id"] = args["categ_id"]
        if "extra_fields" in args:
            values.update(args["extra_fields"])
        new_id = client.create("product.template", values)
        if "cost_price" in args:
            variants = client.search("product.product", [["product_tmpl_id", "=", new_id]])
            if variants:
                client.write("product.product", variants, {"standard_price": args["cost_price"]})
        return _ok({
            "id": new_id,
            "url": record_url("product.template", new_id),
            "template": _with_url(client.get_product_template(new_id), "product.template"),
        })

    if name == "product_set_image":
        model = args.get("model", "product.template")
        record_id = args["id"]
        if "image_url" in args:
            b64 = client.encode_image_from_url(args["image_url"])
        elif "image_base64" in args:
            b64 = args["image_base64"]
        else:
            return _err("Provide either image_url or image_base64")
        client.write(model, [record_id], {"image_1920": b64})
        return _ok({"success": True, "model": model, "id": record_id, "url": record_url(model, record_id)})

    if name == "product_update_variant":
        ok = client.write("product.product", [args["variant_id"]], args["values"])
        return _ok({"success": ok, "variant_id": args["variant_id"], "url": record_url("product.product", args["variant_id"])})

    if name == "task_upsert":
        values: dict = {}
        for field in ["name", "project_id", "description", "stage_id",
                      "date_deadline", "priority"]:
            if field in args:
                values[field] = args[field]
        if "user_ids" in args:
            values["user_ids"] = [(6, 0, args["user_ids"])]
        if "tag_ids" in args:
            values["tag_ids"] = [(6, 0, args["tag_ids"])]
        if "extra_fields" in args:
            values.update(args["extra_fields"])

        task_id = args.get("task_id")
        if task_id:
            client.write("project.task", [task_id], values)
            return _ok({"updated": True, "task_id": task_id, "url": record_url("project.task", task_id)})
        else:
            new_id = client.create("project.task", values)
            return _ok({"created": True, "task_id": new_id, "url": record_url("project.task", new_id)})

    if name == "helpdesk_ticket_upsert":
        values: dict = {}
        for field in ["name", "partner_id", "partner_name", "description",
                      "team_id", "user_id", "stage_id", "priority"]:
            if field in args:
                values[field] = args[field]
        if "tag_ids" in args:
            values["tag_ids"] = [(6, 0, args["tag_ids"])]
        if "extra_fields" in args:
            values.update(args["extra_fields"])

        ticket_id = args.get("ticket_id")
        if ticket_id:
            client.write("helpdesk.ticket", [ticket_id], values)
            return _ok({"updated": True, "ticket_id": ticket_id, "url": record_url("helpdesk.ticket", ticket_id)})
        else:
            new_id = client.create("helpdesk.ticket", values)
            return _ok({"created": True, "ticket_id": new_id, "url": record_url("helpdesk.ticket", new_id)})

    # ── Datasheet generation ──────────────────────────────────────────────────
    if name == "generate_datasheet":
        import base64 as _b64

        def _to_data_uri(b64_val: str | bool | None) -> str:
            """Convert an Odoo base64 field to a MIME-typed data URI."""
            if not b64_val:
                return ""
            try:
                raw = _b64.b64decode(b64_val)
            except Exception:
                raw = b64_val if isinstance(b64_val, bytes) else b""
            if raw[:4] == b"\x89PNG":
                mime = "image/png"
            elif raw[:2] == b"\xff\xd8":
                mime = "image/jpeg"
            elif raw[:4] == b"RIFF" and raw[8:12] == b"WEBP":
                mime = "image/webp"
            elif raw[:6] in (b"GIF87a", b"GIF89a"):
                mime = "image/gif"
            elif b"<svg" in raw[:256]:
                mime = "image/svg+xml"
            else:
                mime = "image/jpeg"
            return f"data:{mime};base64,{_b64.b64encode(raw).decode('ascii')}"

        # 1. Resolve the variant
        variant_id = args.get("variant_id")
        if not variant_id:
            if args.get("internal_reference"):
                hits = client.search("product.product", [["default_code", "=", args["internal_reference"]]])
            elif args.get("product_name"):
                hits = client.search("product.product", [["name", "ilike", args["product_name"]]])
            else:
                return _err("Provide internal_reference, variant_id, or product_name")
            if not hits:
                return _err("No matching product variant found")
            if len(hits) > 1:
                names = client.read("product.product", hits, fields=["default_code", "name"])
                return _err(f"Multiple variants matched — be more specific: {names}")
            variant_id = hits[0]

        # 2. Fetch all spec fields (variant level)
        SPEC_FIELDS = [
            "id", "name", "default_code", "description_sale",
            "product_tmpl_id", "image_1920",
            "product_template_attribute_value_ids",
            "x_studio_lumens_1",
            "x_studio_ip_rating", "x_studio_ik_rating",
            "x_studio_light_source", "x_studio_lifetime",
            "x_studio_sdcm_1", "x_studio_input_voltage",
            "x_studio_dimming_resolution", "x_studio_wiring",
            "x_studio_power_factor", "x_studio_mounting",
            "x_studio_length_mm_1", "x_studio_width_mm_1",
            "x_studio_height_mm_2", "x_studio_diameter_mm_1",
            "x_studio_cut_out", "x_studio_weight_kg",
            "x_studio_specsheet_notes",
            "x_studio_ies_image",
            "x_studio_dimension_image",
            "x_studio_colour_detail",
        ]
        variants = client.read("product.product", [variant_id], fields=SPEC_FIELDS)
        if not variants:
            return _err(f"Variant ID {variant_id} not found")
        v = variants[0]

        # 2b. Fetch template-level description override (x_datasheet_description, only if module installed)
        tmpl_id = v["product_tmpl_id"][0] if isinstance(v.get("product_tmpl_id"), list) else v.get("product_tmpl_id")
        datasheet_description = ""
        if tmpl_id:
            try:
                tmpl_records = client.read("product.template", [tmpl_id], fields=["x_datasheet_description", "description_sale"])
                if tmpl_records:
                    t = tmpl_records[0]
                    datasheet_description = t.get("x_datasheet_description") or t.get("description_sale") or ""
            except Exception:
                # x_datasheet_description only exists when vl_datasheet_pdfmonkey module is installed
                pass

        def _f(val) -> str:
            if val is False or val is None:
                return ""
            return str(val)

        # 3. Fetch and sort attribute values; handle colour+detail
        attr_val_ids = v.get("product_template_attribute_value_ids", [])
        attr_vals = []
        known_attrs: dict[str, str] = {}
        colour_detail = _f(v.get("x_studio_colour_detail"))
        if attr_val_ids:
            av_records = client.read(
                "product.template.attribute.value",
                attr_val_ids,
                fields=["attribute_id", "name"],
            )
            for av in av_records:
                attr_name = av["attribute_id"][1] if isinstance(av["attribute_id"], list) else str(av["attribute_id"])
                attr_value = av["name"]
                # Append colour detail to colour attribute values
                if attr_name.lower() in ("colour", "color") and colour_detail:
                    attr_value = f"{attr_value} {colour_detail}"
                attr_vals.append({"name": attr_name, "value": attr_value})
                known_attrs[attr_name.lower()] = attr_value

        # 4. Build the payload (matches template variable structure)
        description = _f(v.get("description_sale") or datasheet_description)
        if datasheet_description:
            description = datasheet_description  # template-level override takes priority

        payload = {
            "default_code":        _f(v.get("default_code")),
            "name":                _f(v.get("name")),
            "power":               known_attrs.get("power", ""),
            "cct_cri":             known_attrs.get("cct/cri", ""),
            "lumens":              _f(v.get("x_studio_lumens_1")),
            "optic":               known_attrs.get("optic", ""),
            "dimming":             known_attrs.get("dimming/control", ""),
            "description_sale":    description,
            "lumen_output":        _f(v.get("x_studio_lumens_1")),
            "ip_rating":           _f(v.get("x_studio_ip_rating")),
            "ik_rating":           _f(v.get("x_studio_ik_rating")),
            "light_source":        _f(v.get("x_studio_light_source")),
            "lifetime":            _f(v.get("x_studio_lifetime")),
            "sdcm":                _f(v.get("x_studio_sdcm_1")),
            "input_voltage":       _f(v.get("x_studio_input_voltage")),
            "dimming_resolution":  _f(v.get("x_studio_dimming_resolution")),
            "wiring":              _f(v.get("x_studio_wiring")),
            "power_factor":        _f(v.get("x_studio_power_factor")),
            "mounting":            _f(v.get("x_studio_mounting")),
            "length_mm":           _f(v.get("x_studio_length_mm_1")),
            "width_mm":            _f(v.get("x_studio_width_mm_1")),
            "height_mm":           _f(v.get("x_studio_height_mm_2")),
            "diameter_mm":         _f(v.get("x_studio_diameter_mm_1")),
            "cutout":              _f(v.get("x_studio_cut_out")),
            "weight":              _f(v.get("x_studio_weight_kg")),
            "notes":               _f(v.get("x_studio_specsheet_notes")),
            "attributes":          attr_vals,
            "product_image":       _to_data_uri(v.get("image_1920")),
            "ies_image":           _to_data_uri(v.get("x_studio_ies_image")),
            "dimension_image":     _to_data_uri(v.get("x_studio_dimension_image")),
            "quote_name":          "",
            "project_legend":      "",
        }

        # 5. Submit to PDFMonkey and return result
        result = generate_document(payload, template_id=args.get("template_id"))
        result["variant_id"] = variant_id
        result["product"] = _f(v.get("name"))
        result["sku"] = _f(v.get("default_code"))
        return _ok(result)

    # ── Staging-only tools ────────────────────────────────────────────────────
    if name == "odoo_call":
        result = client.call(args["model"], args["method"], args["ids"], **args.get("kwargs", {}))
        return _ok(result)

    if name == "delivery_split":
        picking_id = args["picking_id"]
        batches = args["batches"]
        if len(batches) < 2:
            return _err("Need at least 2 batches to perform a split.")

        created: list[dict] = []
        remaining = picking_id
        for i, batch in enumerate(batches[:-1]):
            new_ids = client.split_picking(remaining, batch["moves"])
            if batch.get("scheduled_date"):
                client.write("stock.picking", [remaining], {"scheduled_date": batch["scheduled_date"]})
            created.append({"batch": i + 1, "picking_id": remaining, "scheduled_date": batch.get("scheduled_date"), "url": record_url("stock.picking", remaining)})
            if new_ids:
                remaining = new_ids[0]

        last = batches[-1]
        if last.get("scheduled_date"):
            client.write("stock.picking", [remaining], {"scheduled_date": last["scheduled_date"]})
        created.append({"batch": len(batches), "picking_id": remaining, "scheduled_date": last.get("scheduled_date"), "url": record_url("stock.picking", remaining)})
        return _ok({"split_pickings": created})

    if name == "delivery_split_from_excel":
        import openpyxl
        from datetime import datetime
        from collections import defaultdict

        picking_id = args["picking_id"]
        wb = openpyxl.load_workbook(args["excel_path"], data_only=True)
        ws = wb[args["sheet_name"]] if args.get("sheet_name") else wb.active
        product_col = args.get("product_col", "Product")
        qty_col = args.get("qty_col", "Qty")
        date_col = args.get("date_col", "Date")

        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            prod_idx, qty_idx, date_idx = headers.index(product_col), headers.index(qty_col), headers.index(date_col)
        except ValueError as e:
            return _err(f"Column not found: {e}. Available: {headers}")

        moves = client.search_read(
            "stock.move",
            [["picking_id", "=", picking_id], ["state", "not in", ["done", "cancel"]]],
            fields=["id", "product_id", "product_uom_qty", "name"],
        )

        def match_move(product_name: str) -> int | None:
            pname = product_name.lower()
            for m in moves:
                pid_name = (m["product_id"][1] if isinstance(m["product_id"], list) else "").lower()
                if pname in pid_name or pid_name in pname:
                    return m["id"]
            prods = client.search_read(
                "product.product",
                ["|", ["name", "ilike", product_name], ["default_code", "ilike", product_name]],
                fields=["id"], limit=1,
            )
            if prods:
                pid = prods[0]["id"]
                for m in moves:
                    mid_pid = m["product_id"][0] if isinstance(m["product_id"], list) else None
                    if mid_pid == pid:
                        return m["id"]
            return None

        date_groups: dict = defaultdict(list)
        unmatched = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            prod_val, qty_val, date_val = row[prod_idx], row[qty_idx], row[date_idx]
            if not prod_val or not qty_val:
                continue
            move_id = match_move(str(prod_val).strip())
            if move_id is None:
                unmatched.append(str(prod_val))
                continue
            date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val).strip() if date_val else None
            date_groups[date_str].append({"move_id": move_id, "qty": float(qty_val)})

        if unmatched:
            return _err(f"Could not match products to moves: {unmatched}")

        batches = [{"scheduled_date": d, "moves": date_groups[d]} for d in sorted(date_groups)]
        if len(batches) < 2:
            return _err("Schedule only has one date — no split needed.")

        return await _dispatch("delivery_split", {"picking_id": picking_id, "batches": batches})

    return _err(f"Unknown tool: {name}")


# ── Entry point ───────────────────────────────────────────────────────────────

async def main():
    async with stdio_server() as (read_stream, write_stream):
        await app.run(read_stream, write_stream, app.create_initialization_options())


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
